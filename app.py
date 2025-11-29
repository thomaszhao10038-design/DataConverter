import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side, numbers
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils import get_column_letter

# --- Configuration ---
POWER_COL_OUT = 'PSumW'
ENERGY_COL = 'Energy (kWh)'
CUMULATIVE_ENERGY_COL = 'Cumulative Energy'

# -----------------------------
# PROCESS SINGLE SHEET
# -----------------------------
def process_sheet(df, date_col, time_col, psum_col):
    """
    Processes a single DataFrame sheet: cleans data, rounds timestamps to 10-minute intervals,
    filters out leading/trailing zero periods, and prepares data for Excel output.
    Periods outside the first and last non-zero reading are filled with NaN (blank) upon re-indexing.
    
    This version combines separate Date and Time columns into a single timestamp index and
    calculates energy consumption.
    """
    df.columns = df.columns.astype(str).str.strip()
    
    # 1. Combine Date and Time columns into a single timestamp string/series
    combined_dt_series = df[date_col].astype(str) + ' ' + df[time_col].astype(str)
    df['Timestamp'] = pd.to_datetime(combined_dt_series, errors="coerce", dayfirst=True)
    timestamp_col = 'Timestamp'
    
    # 2. Clean and convert power column (handle commas as decimal separators)
    power_series = df[psum_col].astype(str).str.strip()
    power_series = power_series.str.replace(',', '.', regex=False)
    df[psum_col] = pd.to_numeric(power_series, errors='coerce')
    
    # 3. Drop rows where we couldn't parse the timestamp or power value
    df = df.dropna(subset=[timestamp_col, psum_col])
    if df.empty:
        return pd.DataFrame()
    
    # --- CORE LOGIC: FILTER LEADING AND TRAILING ZEROS ---
    non_zero_indices = df[df[psum_col].abs() != 0].index
    
    if non_zero_indices.empty:
        return pd.DataFrame() 
        
    first_valid_idx = non_zero_indices.min()
    last_valid_idx = non_zero_indices.max()
    
    df = df.loc[first_valid_idx:last_valid_idx].copy()
    # ----------------------------------------------------
    
    # Resample data to 10-minute intervals
    df_indexed = df.set_index(timestamp_col)
    df_indexed.index = df_indexed.index.floor('10min')
    # Sum the power values within each 10-minute slot
    resampled_data = df_indexed[psum_col].groupby(level=0).sum()
    
    df_out = resampled_data.reset_index()
    df_out.columns = ['Rounded', POWER_COL_OUT] 
    
    if df_out.empty or df_out['Rounded'].isna().all():
        return pd.DataFrame()
    
    # Get the original dates present in the processed data
    original_dates = set(df_out['Rounded'].dt.date)
    
    # Create a full 10-minute index from the start of the first day to the end of the last day
    min_dt = df_out['Rounded'].min().floor('D')
    max_dt_exclusive = df_out['Rounded'].max().ceil('D') 
    if min_dt >= max_dt_exclusive:
        return pd.DataFrame()
    
    full_time_index = pd.date_range(
        start=min_dt.to_pydatetime(),
        end=max_dt_exclusive.to_pydatetime(),
        freq='10min',
        inclusive='left'
    )
    
    # Reindex with the full index, filling missing slots with NaN (blank)
    df_indexed_for_reindex = df_out.set_index('Rounded')
    df_padded_series = df_indexed_for_reindex[POWER_COL_OUT].reindex(full_time_index) 
    
    grouped = df_padded_series.reset_index().rename(columns={'index': 'Rounded'})
    grouped.columns = ['Rounded', POWER_COL_OUT]
    
    # Ensure the column is float type to correctly hold NaN values
    grouped[POWER_COL_OUT] = grouped[POWER_COL_OUT].astype(float) 

    grouped["Date"] = grouped["Rounded"].dt.date
    grouped["Time"] = grouped["Rounded"].dt.strftime("%H:%M") 
    
    # Filter back to only the dates originally present in the file
    grouped = grouped[grouped["Date"].isin(original_dates)]
    
    # Add kW column (absolute value)
    grouped['kW'] = grouped[POWER_COL_OUT].abs() / 1000
    
    # --- NEW: ENERGY CALCULATION (Energy = Power * Time Interval) ---
    # 10 minutes is 10/60 = 1/6 hour. Energy (kWh) = Power (kW) * Time (h)
    # Power (W) to Power (kW): / 1000
    # Time (10 min) to Time (h): / 6
    # Energy (kWh) = |PSumW| / 1000 * (1/6) = |PSumW| / 6000
    grouped[ENERGY_COL] = grouped[POWER_COL_OUT].abs() / 6000
    
    # --- NEW: CUMULATIVE ENERGY ---
    # Calculate cumulative energy group by day, excluding NaNs from the calculation
    grouped[CUMULATIVE_ENERGY_COL] = grouped.groupby("Date")[ENERGY_COL].transform(lambda x: x.cumsum().where(x.notna()))
    
    return grouped

# -----------------------------
# BUILD EXCEL
# -----------------------------

def get_col_names_for_chart(col_start):
    """Helper to get excel column letters for chart data (kW, kWh, Cumulative)"""
    # Assuming columns are: UTC (1), Time (2), W (3), kW (4), Energy (5), Cumulative (6)
    # The chart columns are kW (col_start + 3), Energy (col_start + 4), Cumulative (col_start + 5)
    return [col_start + 3, col_start + 4, col_start + 5]

def create_individual_sheet_charts(ws, dates, day_intervals, max_row_used, col_data_map, chart_type):
    """Generates a line chart for the specified column (kW or Cumulative Energy)."""
    
    if not dates:
        return
    
    chart = LineChart()
    if chart_type == 'kW':
        title = "Daily 10-Minute Absolute Power Profile"
        y_title = "kW"
        col_offset = 3 # kW column index relative to col_start
        chart_pos = f'G{max_row_used+2}'
        chart.height = 12.5
        chart.width = 23    
    elif chart_type == 'Cumulative':
        title = "Daily Accumulated Energy Profile"
        y_title = "Cumulative Energy (kWh)"
        col_offset = 5 # Cumulative Energy column index relative to col_start
        chart_pos = f'G{max_row_used+20}'
        chart.height = 12.5 
        chart.width = 23    
    else:
        return

    chart.title = f"{title} - {ws.title}"
    chart.y_axis.title = y_title
    chart.x_axis.title = "Time"
    
    max_rows = max(day_intervals)
    first_time_col = 2
    categories_ref = Reference(ws, min_col=first_time_col, min_row=3, max_row=2 + max_rows)

    col_start = 1
    for i, n_rows in enumerate(day_intervals):
        # Data ref: starts at row 3, ends at 2+n_rows
        data_ref = Reference(ws, min_col=col_start + col_offset, min_row=3, 
                                max_col=col_start + col_offset, max_row=2 + n_rows)
        
        date_title_str = dates[i].strftime('%d-%b')
        
        s = Series(values=data_ref, title=date_title_str)
        chart.series.append(s)
        
        col_start += 6 # Each day now uses 6 columns (UTC, Time, W, kW, Energy, Cumulative)

    chart.set_categories(categories_ref)
    ws.add_chart(chart, chart_pos)

def build_output_excel(sheets_dict):
    """Creates the final formatted Excel file with data, charts, and summary."""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    title_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))

    # Data structure for the "Total" sheet
    # Format: { date_obj: { sheet_name: {'max_kw': val, 'avg_kw': val, 'total_kwh': val}, ... }, ... }
    total_sheet_data = {}
    sheet_names_list = []

    for sheet_name, df in sheets_dict.items():
        ws = wb.create_sheet(sheet_name)
        sheet_names_list.append(sheet_name)
        dates = sorted(df["Date"].unique())
        col_start = 1
        max_row_used = 0
        daily_summary_list = [] # Stores (date_short, max_kw, avg_kw, total_kwh)
        day_intervals = []
        
        # Structure (New):
        # Row 1: Merged Date Header (Full Date)
        # Row 2: Sub-headers (UTC, Local Time Stamp, Active Power (W), kW, Energy (kWh), Cumulative Energy)
        # Row 3: Start of data 

        for date in dates:
            # Get all data for the day (including NaNs for missing periods)
            day_data_full = df[df["Date"] == date].sort_values("Time")
            
            # Data used for calculations (excluding the new NaNs from outside the active period)
            day_data_active = day_data_full.dropna(subset=[POWER_COL_OUT])
            
            n_rows = len(day_data_full) # Use full count for row structure
            day_intervals.append(n_rows)
            
            data_start_row = 3 # Data starts at Row 3 
            merge_start = data_start_row
            merge_end = merge_start + n_rows - 1

            date_str_full = date.strftime('%Y-%m-%d')
            date_str_short = date.strftime('%d-%b')

            # Row 1: Merge date header (Long Date) - spans 6 columns
            ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start+5)
            ws.cell(row=1, column=col_start, value=date_str_full).alignment = Alignment(horizontal="center")
            
            # Row 2: Sub-headers (6 columns)
            ws.cell(row=2, column=col_start, value="UTC Offset (minutes)")
            ws.cell(row=2, column=col_start+1, value="Local Time Stamp")
            ws.cell(row=2, column=col_start+2, value="Active Power (W)")
            ws.cell(row=2, column=col_start+3, value="kW")
            ws.cell(row=2, column=col_start+4, value="Energy (kWh)") # NEW
            ws.cell(row=2, column=col_start+5, value="Cumulative Energy") # NEW

            # Merge UTC column (Starts at row 3)
            ws.merge_cells(start_row=merge_start, start_column=col_start, end_row=merge_end, end_column=col_start)
            ws.cell(row=merge_start, column=col_start, value=date_str_full).alignment = Alignment(horizontal="center", vertical="center")

            # Fill data (starts at row 3)
            for idx, r in enumerate(day_data_full.itertuples(), start=merge_start):
                ws.cell(row=idx, column=col_start+1, value=r.Time)
                ws.cell(row=idx, column=col_start+2, value=getattr(r, POWER_COL_OUT)) 
                ws.cell(row=idx, column=col_start+3, value=r.kW)
                ws.cell(row=idx, column=col_start+4, value=getattr(r, ENERGY_COL)).number_format = numbers.FORMAT_NUMBER_00 # NEW
                ws.cell(row=idx, column=col_start+5, value=getattr(r, CUMULATIVE_ENERGY_COL)).number_format = numbers.FORMAT_NUMBER_00 # NEW

            # Summary stats
            stats_row_start = merge_end + 1
            
            # W Stats
            sum_w = day_data_active[POWER_COL_OUT].sum()
            mean_w = day_data_active[POWER_COL_OUT].mean()
            max_w = day_data_active[POWER_COL_OUT].max()
            
            # kW Stats
            sum_kw = day_data_active['kW'].sum()
            mean_kw = day_data_active['kW'].mean()
            max_kw = day_data_active['kW'].max()
            
            # Energy Stats (kWh)
            total_kwh = day_data_active[ENERGY_COL].sum()
            # Daily Average Energy is Total kWh / (Number of 10-min slots in 24h) * (10 min in hours) = Total kWh / 24h
            # Average power over 24 hours (kW) = Total Energy (kWh) / 24 hours.
            avg_kw_24h = total_kwh / 24.0 # NEW

            # --- Write Stats (Starting from Col+1: Time) ---
            
            # Total
            ws.cell(row=stats_row_start, column=col_start+1, value="Total")
            ws.cell(row=stats_row_start, column=col_start+2, value=sum_w).number_format = numbers.FORMAT_NUMBER_00
            ws.cell(row=stats_row_start, column=col_start+3, value=sum_kw).number_format = numbers.FORMAT_NUMBER_00
            ws.cell(row=stats_row_start, column=col_start+4, value=total_kwh).number_format = numbers.FORMAT_NUMBER_00 # NEW: Total Energy

            # Average (Active Period)
            ws.cell(row=stats_row_start+1, column=col_start+1, value="Avg (Active W/kW)")
            ws.cell(row=stats_row_start+1, column=col_start+2, value=mean_w).number_format = numbers.FORMAT_NUMBER_00
            ws.cell(row=stats_row_start+1, column=col_start+3, value=mean_kw).number_format = numbers.FORMAT_NUMBER_00
            
            # Max
            ws.cell(row=stats_row_start+2, column=col_start+1, value="Max")
            ws.cell(row=stats_row_start+2, column=col_start+2, value=max_w).number_format = numbers.FORMAT_NUMBER_00
            ws.cell(row=stats_row_start+2, column=col_start+3, value=max_kw).number_format = numbers.FORMAT_NUMBER_00

            # NEW: Daily Average Power (24h) and Energy Summary Row
            ws.cell(row=stats_row_start+3, column=col_start+1, value="Avg (24h kW)") # NEW
            ws.cell(row=stats_row_start+3, column=col_start+3, value=avg_kw_24h).number_format = numbers.FORMAT_NUMBER_00 # NEW

            max_row_used = max(max_row_used, stats_row_start+3)
            # Store Max kW, 24h Avg kW, Total kWh for Total sheet and summary table
            daily_summary_list.append((date_str_short, max_kw, avg_kw_24h, total_kwh)) 

            # Collect data for "Total" sheet
            if date not in total_sheet_data:
                total_sheet_data[date] = {}
            total_sheet_data[date][sheet_name] = {'max_kw': max_kw, 'avg_kw': avg_kw_24h, 'total_kwh': total_kwh}

            col_start += 6 # Each day now takes up 6 columns

        # Add Line Charts for Individual Sheet
        create_individual_sheet_charts(ws, dates, day_intervals, max_row_used, get_col_names_for_chart(col_start), 'kW')
        create_individual_sheet_charts(ws, dates, day_intervals, max_row_used, get_col_names_for_chart(col_start), 'Cumulative')


        # Add Daily Summary Table (Max kW, Avg kW, Total kWh)
        if daily_summary_list:
            start_row = max_row_used + 5 
            
            ws.cell(row=start_row, column=1, value="Daily Performance Summary").font = title_font
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
            start_row += 1

            ws.cell(row=start_row, column=1, value="Day").fill = header_fill
            ws.cell(row=start_row, column=2, value="Max kW").fill = header_fill
            ws.cell(row=start_row, column=3, value="Avg kW (24h)").fill = header_fill # NEW
            ws.cell(row=start_row, column=4, value="Total kWh").fill = header_fill # NEW

            for d, (date_str, max_kw, avg_kw, total_kwh) in enumerate(daily_summary_list):
                row = start_row+1+d
                ws.cell(row=row, column=1, value=date_str)
                ws.cell(row=row, column=2, value=max_kw).number_format = numbers.FORMAT_NUMBER_00
                ws.cell(row=row, column=3, value=avg_kw).number_format = numbers.FORMAT_NUMBER_00
                ws.cell(row=row, column=4, value=total_kwh).number_format = numbers.FORMAT_NUMBER_00

            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15 # NEW

    # -----------------------------
    # CREATE MODIFIED "TOTAL" SHEET
    # -----------------------------
    if total_sheet_data:
        ws_total = wb.create_sheet("Total")
        sorted_dates = sorted(total_sheet_data.keys())
        data_max_row = len(sorted_dates)
        
        # --- Prepare Headers and Initial Setup ---
        
        # New structure: Date in Col A, then groups of 3 columns (Max, Avg, Total) for each sheet, then Total Load (3 cols)
        
        # Header Row 1: Merge Sheet Names
        col_idx = 2
        for sheet_name in sheet_names_list:
            ws_total.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+2)
            ws_total.cell(row=1, column=col_idx, value=sheet_name).font = title_font
            col_idx += 3
        
        ws_total.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+2)
        ws_total.cell(row=1, column=col_idx, value="Total Load").font = title_font

        # Header Row 2: Date, Max kW, Avg kW, Total kWh (repeated)
        headers_row_2 = ["Date"]
        for _ in sheet_names_list:
            headers_row_2.extend(["Max kW", "Avg kW (24h)", "Total kWh"])
        headers_row_2.extend(["Max Total kW", "Avg Total kW (24h)", "Total Total kWh"])

        for col_idx, header_text in enumerate(headers_row_2, 1):
            cell = ws_total.cell(row=2, column=col_idx, value=header_text)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws_total.column_dimensions[get_column_letter(col_idx)].width = 16

        # --- Write Data (Date Rows) ---
        max_total_kw_list = []
        avg_total_kw_list = []
        total_total_kwh_list = []

        for row_idx, date_obj in enumerate(sorted_dates, 3): # Data starts at Row 3
            # Date Column
            date_str = date_obj.strftime('%Y-%m-%d')
            date_cell = ws_total.cell(row=row_idx, column=1, value=date_str)
            date_cell.border = thin_border
            date_cell.alignment = Alignment(horizontal="center")
            
            daily_max_kw_sum = 0
            daily_avg_kw_sum = 0
            daily_total_kwh_sum = 0
            
            col_idx = 2
            # Sheet Columns (Max kW, Avg kW, Total kWh)
            for sheet_name in sheet_names_list:
                data = total_sheet_data[date_obj].get(sheet_name, {'max_kw': 0, 'avg_kw': 0, 'total_kwh': 0})
                
                max_kw = data['max_kw'] if not pd.isna(data['max_kw']) else 0
                avg_kw = data['avg_kw'] if not pd.isna(data['avg_kw']) else 0
                total_kwh = data['total_kwh'] if not pd.isna(data['total_kwh']) else 0
                
                # Max kW
                cell = ws_total.cell(row=row_idx, column=col_idx, value=max_kw)
                cell.number_format = numbers.FORMAT_NUMBER_00
                cell.border = thin_border
                col_idx += 1
                daily_max_kw_sum += max_kw
                
                # Avg kW (24h)
                cell = ws_total.cell(row=row_idx, column=col_idx, value=avg_kw)
                cell.number_format = numbers.FORMAT_NUMBER_00
                cell.border = thin_border
                col_idx += 1
                daily_avg_kw_sum += avg_kw
                
                # Total kWh
                cell = ws_total.cell(row=row_idx, column=col_idx, value=total_kwh)
                cell.number_format = numbers.FORMAT_NUMBER_00
                cell.border = thin_border
                col_idx += 1
                daily_total_kwh_sum += total_kwh

            # Total Load Columns
            max_total_kw_list.append(daily_max_kw_sum)
            avg_total_kw_list.append(daily_avg_kw_sum)
            total_total_kwh_list.append(daily_total_kwh_sum)

            # Max Total kW
            total_max_cell = ws_total.cell(row=row_idx, column=col_idx, value=daily_max_kw_sum)
            total_max_cell.number_format = numbers.FORMAT_NUMBER_00
            total_max_cell.border = thin_border
            total_max_cell.font = Font(bold=True)
            col_idx += 1
            
            # Avg Total kW
            total_avg_cell = ws_total.cell(row=row_idx, column=col_idx, value=daily_avg_kw_sum)
            total_avg_cell.number_format = numbers.FORMAT_NUMBER_00
            total_avg_cell.border = thin_border
            total_avg_cell.font = Font(bold=True)
            col_idx += 1

            # Total Total kWh
            total_kwh_cell = ws_total.cell(row=row_idx, column=col_idx, value=daily_total_kwh_sum)
            total_kwh_cell.number_format = numbers.FORMAT_NUMBER_00
            total_kwh_cell.border = thin_border
            total_kwh_cell.font = Font(bold=True)
            # col_idx += 1 # Not needed

        # --- Write Summary Rows (Average and Total) - NEW REQUEST ---
        
        # Calculate overall averages and totals
        overall_stats = {
            'max_kw': 0, 'avg_kw': 0, 'total_kwh': 0, 
            'total_max_kw': sum(max_total_kw_list),
            'total_avg_kw': sum(avg_total_kw_list),
            'total_total_kwh': sum(total_total_kwh_list)
        }
        
        # Calculate Total row (Sum of columns)
        total_row = data_max_row + 3 # New Row after data
        ws_total.cell(row=total_row, column=1, value="Total").font = bold_font
        
        col_idx = 2
        for sheet_name in sheet_names_list:
            # Max kW Sum (Overall)
            formula_max = f"=SUM({get_column_letter(col_idx)}{3}:{get_column_letter(col_idx)}{data_max_row+2})"
            cell = ws_total.cell(row=total_row, column=col_idx, value=formula_max)
            cell.number_format = numbers.FORMAT_NUMBER_00
            col_idx += 1
            
            # Avg kW (24h) Sum (Overall) - Summing averages is questionable, but following the row/column structure
            formula_avg = f"=SUM({get_column_letter(col_idx)}{3}:{get_column_letter(col_idx)}{data_max_row+2})"
            cell = ws_total.cell(row=total_row, column=col_idx, value=formula_avg)
            cell.number_format = numbers.FORMAT_NUMBER_00
            col_idx += 1

            # Total kWh Sum (Overall) - This is a valid sum (Grand Total Energy)
            formula_kwh = f"=SUM({get_column_letter(col_idx)}{3}:{get_column_letter(col_idx)}{data_max_row+2})"
            cell = ws_total.cell(row=total_row, column=col_idx, value=formula_kwh)
            cell.number_format = numbers.FORMAT_NUMBER_00
            col_idx += 1
            
        # Total Load Total (sums should be equal to the total lists sum, use the lists for certainty)
        ws_total.cell(row=total_row, column=col_idx, value=overall_stats['total_max_kw']).number_format = numbers.FORMAT_NUMBER_00
        col_idx += 1
        ws_total.cell(row=total_row, column=col_idx, value=overall_stats['total_avg_kw']).number_format = numbers.FORMAT_NUMBER_00
        col_idx += 1
        ws_total.cell(row=total_row, column=col_idx, value=overall_stats['total_total_kwh']).number_format = numbers.FORMAT_NUMBER_00
        # col_idx += 1 # Not needed

        # --- Add Total Charts ---

        # 1st Line Graph: Superimposed Daily Max Power (kW)
        if sorted_dates:
            chart_max_kw = LineChart()
            chart_max_kw.title = "Overview: Daily Max Power (kW)" 
            chart_max_kw.y_axis.title = "Max Power (kW)"
            chart_max_kw.x_axis.title = "Date"
            chart_max_kw.height = 15
            chart_max_kw.width = 30
            
            # Data reference for Max kW columns (Col 2, 5, 8, ... and the final Max Total)
            max_kw_cols = [2] + [i*3 + 2 for i in range(1, len(sheet_names_list))] + [len(headers_row_2)] 
            
            data_ref_max = Reference(ws_total, min_col=2, min_row=2, max_col=len(headers_row_2), max_row=data_max_row+2)
            chart_max_kw.add_data(data_ref_max, titles_from_data=True, exceptedCols=tuple(i for i in range(1, len(headers_row_2)-1) if i % 3 != 0))

            # Manual series creation to ensure only Max kW columns are picked (Col 2, 5, 8, ...)
            max_total_col_idx = len(headers_row_2)
            chart_max_kw = LineChart()
            chart_max_kw.title = "Overview: Daily Max Power (kW)" 
            chart_max_kw.y_axis.title = "Max Power (kW)"
            chart_max_kw.x_axis.title = "Date"
            chart_max_kw.height = 15
            chart_max_kw.width = 30
            
            for i, col in enumerate([2] + [i*3 + 2 for i in range(1, len(sheet_names_list))]):
                s_ref = Reference(ws_total, min_col=col, min_row=3, max_row=data_max_row+2)
                t_ref = Reference(ws_total, min_col=col, min_row=2, max_row=2)
                s = Series(values=s_ref, title_from_data=True)
                chart_max_kw.series.append(s)
            
            # Add Total Load Max kW Series (last column)
            s_ref_total = Reference(ws_total, min_col=max_total_col_idx, min_row=3, max_row=data_max_row+2)
            s_total = Series(values=s_ref_total, title=headers_row_2[-3]) # Use "Max Total kW" as title
            chart_max_kw.series.append(s_total)
            
            for s in chart_max_kw.series:
                s.smooth = False

            cats_ref = Reference(ws_total, min_col=1, min_row=3, max_row=data_max_row+2)
            chart_max_kw.set_categories(cats_ref)
            
            ws_total.add_chart(chart_max_kw, "B" + str(total_row + 3))

            # 2nd Line Graph: Superimposed Accumulated Energy (Total kWh)
            chart_total_kwh = LineChart()
            chart_total_kwh.title = "Overview: Accumulated Energy (Total kWh)" 
            chart_total_kwh.y_axis.title = "Total Energy (kWh)"
            chart_total_kwh.x_axis.title = "Date"
            chart_total_kwh.height = 15
            chart_total_kwh.width = 30
            
            # Data reference for Total kWh columns (Col 4, 7, 10, ... and the final Total Total kWh)
            total_kwh_cols = [4] + [i*3 + 4 for i in range(1, len(sheet_names_list))] + [len(headers_row_2)] 
            total_total_kwh_col_idx = len(headers_row_2)

            for i, col in enumerate([4] + [i*3 + 4 for i in range(1, len(sheet_names_list))]):
                # Use the 'Total' row data (cumulative sum of the columns)
                s_ref = Reference(ws_total, min_col=col, min_row=3, max_row=data_max_row+2)
                t_ref = Reference(ws_total, min_col=col, min_row=2, max_row=2)
                s = Series(values=s_ref, title_from_data=True)
                chart_total_kwh.series.append(s)
            
            # Add Total Load Total kWh Series
            s_ref_total_kwh = Reference(ws_total, min_col=total_total_kwh_col_idx, min_row=3, max_row=data_max_row+2)
            s_total_kwh = Series(values=s_ref_total_kwh, title=headers_row_2[-1]) # Use "Total Total kWh" as title
            chart_total_kwh.series.append(s_total_kwh)

            for s in chart_total_kwh.series:
                s.smooth = False
            
            # Categories are the same Dates
            chart_total_kwh.set_categories(cats_ref)
            
            ws_total.add_chart(chart_total_kwh, "B" + str(total_row + 21))


    stream = BytesIO()
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > len(sheets_dict) + (1 if total_sheet_data else 0):
        wb.remove(wb['Sheet'])
        
    wb.save(stream)
    stream.seek(0)
    return stream

# -----------------------------
# STREAMLIT APP
# -----------------------------
def app():
    st.set_page_config(layout="wide", page_title="Electricity Data Converter")
    st.title("📊 Excel 10-Minute Electricity Data Converter")
    st.markdown("""
        Upload an **Excel file (.xlsx)** with time-series data. Each sheet is processed to calculate total absolute power (W) in 10-minute intervals. 
        
        **Input Format Expected:** Separate columns for **Date**, **Time**, and **PSum (W)**.
        
        **New Feature:** Leading and trailing zero values (representing missing readings) are now filtered out and appear blank, but zero values *within* the active recording period are kept.
        
        The output Excel file includes:
        1. **Individual Sheet Analysis:** Data is expanded to include **Energy (kWh)** and **Cumulative Energy (kWh)**. It includes a **Max Power Summary table** and **two line charts** (one for daily Max Power profiles and one for daily Cumulative Energy).
        2. **Total Summary Sheet:** A comparative table and **two line graphs** (Superimposed Daily Max Power and Superimposed Accumulated Energy) of daily performance across all sheets.
    """)

    uploaded = st.file_uploader("Upload .xlsx file", type=["xlsx"])

    if uploaded:
        xls = pd.ExcelFile(uploaded)
        result_sheets = {}
        st.write("---")

        for sheet_name in xls.sheet_names:
            st.markdown(f"**Processing sheet: `{sheet_name}`**")
            try:
                df = pd.read_excel(uploaded, sheet_name=sheet_name)
            except Exception as e:
                st.error(f"Error reading sheet '{sheet_name}': {e}")
                continue

            df.columns = df.columns.astype(str).str.strip()
            
            # --- COLUMN DETECTION ---
            date_col = next((c for c in df.columns if c in ["Date","DATE","date"]), None)
            time_col = next((c for c in df.columns if c in ["Time","TIME","time"]), None)
            
            if not date_col or not time_col:
                st.error(f"No valid Date and/or Time column in sheet '{sheet_name}' (expected: Date, Time).")
                continue

            psum_col = next((c for c in df.columns if c in ["PSum (W)","Psum (W)","PSum","P (W)","Power"]), None)
            if not psum_col:
                st.error(f"No valid PSum column in sheet '{sheet_name}' (expected: PSum (W), Power, etc.).")
                continue

            # --- FUNCTION CALL ---
            processed = process_sheet(df, date_col, time_col, psum_col)
            
            if not processed.empty:
                result_sheets[sheet_name] = processed
                st.success(f"Sheet '{sheet_name}' processed successfully with {len(processed['Date'].unique())} day(s) of data.")
            else:
                st.warning(f"Sheet '{sheet_name}' had no usable data (or all readings were zero/missing).")
        
        st.write("---")
        if result_sheets:
            st.balloons()
            st.success("All usable sheets processed. Generating Excel output...")
            output_stream = build_output_excel(result_sheets)
            st.download_button(
                label="📥 Download Converted Excel (Converted_Output.xlsx)",
                data=output_stream,
                file_name="Converted_Output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        elif uploaded:
            st.error("No data could be processed from the uploaded file.")

if __name__ == "__main__":
    app()
